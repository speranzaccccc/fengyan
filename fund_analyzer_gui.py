"""
私募基金净值分析工具 - Mac App 版本
使用 Tkinter 构建，可打包为 .app 直接双击运行
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import math
import re
from datetime import datetime, date
from pathlib import Path
import os
import sys


# ============== 配置区 ==============
RF_ANNUAL = 1.82  # 年化无风险利率（%）
SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
BENCHMARK_NAMES = ("中证1000", "中证800")
# ============== 配置区 ==============


def parse_nav(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip().replace(",", ""))
    except ValueError:
        return None


def parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val).strip()[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_percent_input(value):
    text = str(value or "").strip().replace("%", "").replace("％", "").replace(",", "")
    if not text:
        return None
    return float(text)


def calc_return_from_prices(start_price, end_price):
    if start_price is None or end_price is None or start_price <= 0:
        return None
    return (end_price / start_price - 1) * 100


def get_index_on_or_before(data, target_date):
    result = None
    for record in data:
        d = record[0]
        if d <= target_date:
            result = record
        else:
            break
    return result


def get_nav_on_or_before(data, target_date):
    result = None
    for record in data:
        if len(record) < 2:
            continue
        d = record[0]
        if d <= target_date:
            result = record
        else:
            break
    return result


def shift_years(base_date, years):
    try:
        return date(base_date.year - years, base_date.month, base_date.day)
    except ValueError:
        return date(base_date.year - years, 2, 28)


def get_trailing_start_record(data, target_date):
    start_record = get_nav_on_or_before(data, target_date)
    if start_record:
        return start_record, False
    return data[0], True


def normalize_header(value):
    return str(value or "").strip().replace("\n", "").replace(" ", "")


def find_column(header, predicate):
    for idx, title in enumerate(header):
        if predicate(title):
            return idx
    return None


def parse_fund_rows(rows, fallback_name, sheet_name):
    header_row_idx = None
    for i, row in enumerate(rows):
        joined = " ".join(str(c) if c else "" for c in row)
        if "净值日期" in joined and "单位净值" in joined:
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError("未找到包含'净值日期'和'单位净值'的表头行")

    header = [normalize_header(c) for c in rows[header_row_idx]]
    date_col = find_column(header, lambda h: "净值日期" in h)
    nav_col = find_column(header, lambda h: h == "单位净值")
    if nav_col is None:
        nav_col = find_column(header, lambda h: "单位净值" in h and "累计" not in h)
    name_col = find_column(header, lambda h: "产品名称" in h)
    cum_nav_col = find_column(header, lambda h: "累计净值" in h or "累计单位净值" in h or ("累计" in h and "净值" in h))
    if date_col is None or nav_col is None:
        raise ValueError(f"表头中未找到净值日期或单位净值列: {header}")

    fund_name = fallback_name
    if name_col is not None:
        for row in rows[header_row_idx + 1:]:
            if len(row) > name_col and row[name_col]:
                fund_name = str(row[name_col]).strip()
                break

    data = []
    for row in rows[header_row_idx + 1:]:
        if len(row) <= max(date_col, nav_col):
            continue
        d = parse_date(row[date_col])
        nav = parse_nav(row[nav_col])
        cum_nav = None
        if cum_nav_col is not None and len(row) > cum_nav_col:
            cum_nav = parse_nav(row[cum_nav_col])
        if d and nav is not None and nav > 0:
            data.append((d, nav, cum_nav))
    data.sort(key=lambda x: x[0])
    if not data:
        raise ValueError("已找到表头，但未解析出有效净值数据，请确认日期和单位净值列有数值")

    return fund_name, data, sheet_name


def parse_index_rows(rows, fallback_name, sheet_name):
    header_row_idx = None
    for i, row in enumerate(rows):
        header = [normalize_header(c) for c in row]
        has_date = any(("日期" in h or h.lower() in ("date", "tradedate")) for h in header)
        numeric_columns = sum(1 for h in header if h and not ("日期" in h or h.lower() in ("date", "tradedate")))
        if has_date and numeric_columns >= 1:
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError("未找到包含日期和指数数值的表头行")

    header = [normalize_header(c) for c in rows[header_row_idx]]
    date_col = find_column(header, lambda h: "日期" in h or h.lower() in ("date", "tradedate"))
    name_col = find_column(header, lambda h: h in ("指数名称", "指数简称", "证券简称", "名称", "指标名称"))
    close_col = find_column(
        header,
        lambda h: (
            "收盘" in h
            or "收盘价" in h
            or "点位" in h
            or h.lower() in ("close", "closeprice", "pxlast")
        )
    )
    if date_col is None:
        raise ValueError("表头中未找到日期列")

    series_map = {}
    if close_col is not None:
        for row in rows[header_row_idx + 1:]:
            if len(row) <= max(date_col, close_col):
                continue
            d = parse_date(row[date_col])
            close = parse_nav(row[close_col])
            if not d or close is None or close <= 0:
                continue
            series_name = f"{fallback_name}-{sheet_name}"
            if name_col is not None and len(row) > name_col and row[name_col]:
                series_name = str(row[name_col]).strip()
            series_map.setdefault(series_name, {})[d] = close
    else:
        value_cols = []
        for idx, title in enumerate(header):
            if idx == date_col or idx == name_col:
                continue
            if title:
                value_cols.append((idx, title))
        if not value_cols:
            raise ValueError("未找到收盘价列或可识别的指数列")

        for row in rows[header_row_idx + 1:]:
            if len(row) <= date_col:
                continue
            d = parse_date(row[date_col])
            if not d:
                continue
            for col_idx, title in value_cols:
                if len(row) <= col_idx:
                    continue
                close = parse_nav(row[col_idx])
                if close is not None and close > 0:
                    series_map.setdefault(title, {})[d] = close

    parsed = {}
    for name, date_values in series_map.items():
        data = sorted(date_values.items(), key=lambda x: x[0])
        if len(data) >= 2:
            parsed[name] = data
    if not parsed:
        raise ValueError("已找到表头，但未解析出有效指数数据")
    return parsed


def load_index_data(filepath):
    ext = Path(filepath).suffix.lower()
    if ext not in SUPPORTED_EXCEL_EXTENSIONS:
        raise ValueError("当前版本支持 .xlsx/.xlsm 指数文件；如果是 .xls，请先另存为 .xlsx 后再上传")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    errors = []
    index_map = {}
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            try:
                parsed = parse_index_rows(rows, Path(filepath).stem, ws.title)
                index_map.update(parsed)
            except ValueError as exc:
                errors.append(f"{ws.title}: {exc}")
    finally:
        sheet_names = ", ".join(wb.sheetnames)
        wb.close()

    if not index_map:
        raise ValueError(
            "无法读取指数数据。请确认工作表里有日期列和收盘价列，或日期列加多个指数列。"
            f"\n已检查工作表: {sheet_names}"
            f"\n详细原因: {'; '.join(errors)}"
        )
    return index_map


def load_fund_data(filepath):
    ext = Path(filepath).suffix.lower()
    if ext not in SUPPORTED_EXCEL_EXTENSIONS:
        raise ValueError("当前版本支持 .xlsx/.xlsm 文件；如果是 .xls，请先另存为 .xlsx 后再上传")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    errors = []
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            try:
                fund_name, data, _ = parse_fund_rows(rows, Path(filepath).stem, ws.title)
                return fund_name, data
            except ValueError as exc:
                errors.append(f"{ws.title}: {exc}")
    finally:
        wb.close()

    sheet_names = ", ".join(wb.sheetnames)
    raise ValueError(
        "无法读取净值数据。请确认某个工作表里有'净值日期'和'单位净值'两列。"
        f"\n已检查工作表: {sheet_names}"
        f"\n详细原因: {'; '.join(errors)}"
    )


def calc_annualized_volatility(daily_returns):
    if len(daily_returns) < 2:
        return 0.0
    mean = sum(daily_returns) / len(daily_returns)
    variance = sum((r - mean) ** 2 for r in daily_returns) / (len(daily_returns) - 1)
    return math.sqrt(variance) * math.sqrt(252) * 100


def calc_sharpe_ratio(ann_return, ann_volatility, rf):
    if ann_volatility == 0:
        return 0.0
    return (ann_return - rf) / ann_volatility


def calc_sortino_ratio(daily_returns, rf_daily, rf_annual, ann_return):
    if len(daily_returns) < 2:
        return 0.0
    squared_diffs = [min(r - rf_daily, 0) ** 2 for r in daily_returns]
    downside_vol = math.sqrt(sum(squared_diffs) / len(daily_returns)) * math.sqrt(252) * 100
    if downside_vol == 0:
        return 0.0
    return (ann_return - rf_annual) / downside_vol


def calc_calmar_ratio(ann_return, max_drawdown):
    if max_drawdown == 0:
        return 0.0
    return ann_return / abs(max_drawdown)


def calc_metrics(data, start_nav_record=None, rf_annual=RF_ANNUAL):
    if not data:
        return {}

    if start_nav_record:
        first_nav = start_nav_record[1]
        start_date = start_nav_record[0]
        all_data = [start_nav_record] + data
    else:
        first_nav = data[0][1]
        start_date = data[0][0]
        all_data = data

    last_nav = data[-1][1]
    end_date = data[-1][0]
    cum_nav = data[-1][2] if len(data[-1]) > 2 else None
    days = (end_date - start_date).days

    total_return = (last_nav / first_nav - 1) * 100
    ann_return = ((last_nav / first_nav) ** (365.0 / days) - 1) * 100 if days > 0 else 0.0

    peak = first_nav
    max_dd = 0.0
    for record in all_data:
        nav = record[1]
        if nav > peak:
            peak = nav
        dd = (peak - nav) / peak
        if dd > max_dd:
            max_dd = dd

    daily_returns = []
    for i in range(1, len(all_data)):
        prev_nav = all_data[i - 1][1]
        curr_nav = all_data[i][1]
        if prev_nav > 0:
            daily_returns.append(curr_nav / prev_nav - 1)

    ann_vol = calc_annualized_volatility(daily_returns)
    sharpe = calc_sharpe_ratio(ann_return, ann_vol, rf_annual)
    rf_daily = (1 + rf_annual / 100) ** (1.0 / 252) - 1
    sortino = calc_sortino_ratio(daily_returns, rf_daily, rf_annual, ann_return)
    calmar = calc_calmar_ratio(ann_return, -max_dd * 100)

    return {
        "start_date": start_date,
        "end_date": end_date,
        "days": days,
        "start_nav": first_nav,
        "end_nav": last_nav,
        "cum_nav": cum_nav,
        "total_return": total_return,
        "ann_return": ann_return,
        "max_drawdown": -max_dd * 100,
        "ann_volatility": ann_vol,
        "sharpe_ratio": sharpe,
        "sortino_ratio": sortino,
        "calmar_ratio": calmar,
    }


def calc_one_year_return(data):
    if not data:
        return None

    latest_date = data[-1][0]
    one_year_ago = shift_years(latest_date, 1)

    start_record, _ = get_trailing_start_record(data, one_year_ago)

    end_record = data[-1]

    if start_record[0] == end_record[0]:
        return None

    one_year_return = (end_record[1] / start_record[1] - 1) * 100
    return {
        "start_date": start_record[0],
        "end_date": end_record[0],
        "start_nav": start_record[1],
        "end_nav": end_record[1],
        "one_year_return": one_year_return,
        "cum_nav": end_record[2] if len(end_record) > 2 else None
    }


def build_trailing_report(fund_name, data, rf_annual):
    results = []
    if not data:
        return results

    latest_date = data[-1][0]
    for years in (1, 2, 3):
        target_date = shift_years(latest_date, years)
        start_record, insufficient = get_trailing_start_record(data, target_date)
        if start_record[0] == latest_date:
            continue

        period_data = [record for record in data if start_record[0] < record[0] <= latest_date]
        if not period_data:
            continue

        m = calc_metrics(period_data, start_nav_record=start_record, rf_annual=rf_annual)
        base_label = "近一年" if years == 1 else f"近{years}年"
        m["label"] = f"{base_label}(不足{years}年)" if insufficient else base_label
        m["fund"] = fund_name
        m["note"] = f"数据不足{years}年，按可用区间计算" if insufficient else ""
        results.append(m)

    return results


def build_report(fund_name, data, rf_annual, use_prev_year_end=True):
    results = build_trailing_report(fund_name, data, rf_annual)
    year = data[-1][0].year

    for y in range(year, 2020, -1):
        if use_prev_year_end:
            prev_dec31 = get_nav_on_or_before(data, date(y - 1, 12, 31))
        else:
            prev_dec31 = None

        period_data = [(d, n, c) for d, n, c in data if date(y - 1, 12, 31) < d <= date(y, 12, 31)]
        if not period_data:
            continue

        if prev_dec31 and use_prev_year_end:
            m = calc_metrics(period_data, start_nav_record=prev_dec31, rf_annual=rf_annual)
        else:
            m = calc_metrics(period_data, rf_annual=rf_annual)

        label = f"{y}年初至今" if y == year else f"{y}年"
        m["label"] = label
        m["fund"] = fund_name
        m["note"] = ""
        results.append(m)

    m = calc_metrics(data, rf_annual=rf_annual)
    m["label"] = "成立以来"
    m["fund"] = fund_name
    m["note"] = ""
    results.append(m)

    return results


def autosize_worksheet(ws):
    for column_cells in ws.columns:
        column_letter = openpyxl.utils.get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 30)


def save_results_workbook(filepath, fund_name, data, results, one_year_data):
    latest = data[-1]
    wb = openpyxl.Workbook()

    summary_ws = wb.active
    summary_ws.title = "汇总"
    summary_ws.append(["项目", "值"])
    summary_rows = [
        ["基金名称", fund_name],
        ["最新净值日期", str(latest[0])],
        ["单位净值", f"{latest[1]:.4f}"],
        ["累计净值", f"{latest[2]:.4f}" if latest[2] is not None else "无"],
        ["近一年收益率", f"{one_year_data['one_year_return']:+.2f}%" if one_year_data else "数据不足"],
        ["近一年起始日期", str(one_year_data["start_date"]) if one_year_data else "--"],
        ["近一年起始净值", f"{one_year_data['start_nav']:.4f}" if one_year_data else "--"],
    ]
    for row in summary_rows:
        summary_ws.append(row)

    metrics_ws = wb.create_sheet("业绩指标")
    headers = ["基金", "区间", "起始日期", "结束日期", "天数", "期初净值", "期末净值", "期末累计净值",
               "收益率%", "年化收益%", "最大回撤%",
               "年化波动率%", "Sharpe", "Sortino", "Calmar", "备注"]
    metrics_ws.append(headers)
    for m in results:
        metrics_ws.append([
            m["fund"],
            m["label"],
            str(m["start_date"]),
            str(m["end_date"]),
            m["days"],
            m["start_nav"],
            m["end_nav"],
            m["cum_nav"],
            m["total_return"],
            m["ann_return"],
            m["max_drawdown"],
            m["ann_volatility"],
            m["sharpe_ratio"],
            m["sortino_ratio"],
            m["calmar_ratio"],
            m.get("note", ""),
        ])

    header_fill = openpyxl.styles.PatternFill("solid", fgColor="D9EAF7")
    header_font = openpyxl.styles.Font(bold=True)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        autosize_worksheet(ws)

    for row in metrics_ws.iter_rows(min_row=2, min_col=6, max_col=8):
        for cell in row:
            cell.number_format = "0.0000"
    for row in metrics_ws.iter_rows(min_row=2, min_col=9, max_col=12):
        for cell in row:
            cell.number_format = "0.00"
    for row in metrics_ws.iter_rows(min_row=2, min_col=13, max_col=15):
        for cell in row:
            cell.number_format = "0.0000"

    wb.save(filepath)


def find_result(results, label_prefix):
    for item in results:
        if item.get("label", "").startswith(label_prefix):
            return item
    return None


def get_product_compare_values(product):
    data = product["data"]
    latest = data[-1]
    results = product.get("results", [])
    near_one = find_result(results, "近一年")
    near_two = find_result(results, "近2年")
    near_three = find_result(results, "近3年")
    since = find_result(results, "成立以来")
    notes = [m.get("note", "") for m in (near_one, near_two, near_three) if m and m.get("note")]

    return {
        "fund": product["fund_name"],
        "file": product["filename"],
        "start_date": data[0][0],
        "latest_date": latest[0],
        "records": len(data),
        "latest_nav": latest[1],
        "latest_cum_nav": latest[2] if len(latest) > 2 else None,
        "near_one_return": near_one["total_return"] if near_one else None,
        "near_two_return": near_two["total_return"] if near_two else None,
        "near_three_return": near_three["total_return"] if near_three else None,
        "since_return": since["total_return"] if since else None,
        "since_max_drawdown": since["max_drawdown"] if since else None,
        "since_sharpe": since["sharpe_ratio"] if since else None,
        "note": "；".join(notes),
    }


def save_all_products_workbook(filepath, products, weekly_rows=None, benchmark_rows=None):
    wb = openpyxl.Workbook()
    extra_header_rows = []
    summary_ws = wb.active
    summary_ws.title = "产品对比"
    summary_headers = [
        "产品", "文件", "数据起点", "最新日期", "记录数", "最新单位净值", "最新累计净值",
        "近一年收益%", "近2年收益%", "近3年收益%", "成立以来收益%",
        "成立以来最大回撤%", "成立以来Sharpe", "备注"
    ]
    summary_ws.append(summary_headers)

    for product in products:
        row = get_product_compare_values(product)
        summary_ws.append([
            row["fund"],
            row["file"],
            str(row["start_date"]),
            str(row["latest_date"]),
            row["records"],
            row["latest_nav"],
            row["latest_cum_nav"],
            row["near_one_return"],
            row["near_two_return"],
            row["near_three_return"],
            row["since_return"],
            row["since_max_drawdown"],
            row["since_sharpe"],
            row["note"],
        ])

    metrics_ws = wb.create_sheet("业绩指标")
    metrics_headers = [
        "产品", "文件", "区间", "起始日期", "结束日期", "天数", "期初净值", "期末净值",
        "期末累计净值", "收益率%", "年化收益%", "最大回撤%", "年化波动率%",
        "Sharpe", "Sortino", "Calmar", "备注"
    ]
    metrics_ws.append(metrics_headers)
    for product in products:
        for m in product.get("results", []):
            metrics_ws.append([
                product["fund_name"],
                product["filename"],
                m["label"],
                str(m["start_date"]),
                str(m["end_date"]),
                m["days"],
                m["start_nav"],
                m["end_nav"],
                m["cum_nav"],
                m["total_return"],
                m["ann_return"],
                m["max_drawdown"],
                m["ann_volatility"],
                m["sharpe_ratio"],
                m["sortino_ratio"],
                m["calmar_ratio"],
                m.get("note", ""),
            ])

    weekly_ws = None
    benchmark_header_row = None
    if weekly_rows is not None:
        weekly_ws = wb.create_sheet("周报涨幅")
        weekly_headers = [
            "产品", "单位净值日期", "单位净值", "上周收盘日期", "上周收盘单位净值",
            "本周收盘日期", "本周收盘单位净值", "累计净值日期", "累计净值",
            "本周涨幅%", "区间收益率%", "是否跑赢对标指数", "备注"
        ]
        weekly_ws.append(weekly_headers)
        for row in weekly_rows:
            weekly_ws.append([
                row["fund"],
                row["unit_date"],
                row["unit_nav"],
                row["prev_date"],
                row["prev_nav"],
                row["current_date"],
                row["current_nav"],
                row["cum_date"],
                row["cum_nav"],
                row["weekly_return"],
                row["interval_return"],
                row["benchmark_result"],
                row["note"],
            ])

        weekly_ws.append([])
        weekly_ws.append(["对标指数数据"])
        benchmark_header_row = weekly_ws.max_row + 1
        weekly_ws.append(["指数", "本周区间", "区间起点收盘", "上周收盘", "本周收盘", "本周涨幅%", "区间", "区间涨幅%"])
        for row in benchmark_rows or []:
            weekly_ws.append([
                row["name"],
                row["weekly_period"],
                row["start_close"],
                row["prev_close"],
                row["current_close"],
                row["weekly_return"],
                row["interval_period"],
                row["interval_return"],
            ])
        extra_header_rows.append((weekly_ws, benchmark_header_row))

    header_fill = openpyxl.styles.PatternFill("solid", fgColor="D9EAF7")
    header_font = openpyxl.styles.Font(bold=True)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        autosize_worksheet(ws)
    for ws, row_idx in extra_header_rows:
        for cell in ws[row_idx]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    for row in summary_ws.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = "0.0000"
    for row in summary_ws.iter_rows(min_row=2, min_col=8, max_col=12):
        for cell in row:
            cell.number_format = "0.00"
    for row in summary_ws.iter_rows(min_row=2, min_col=13, max_col=13):
        for cell in row:
            cell.number_format = "0.0000"

    for row in metrics_ws.iter_rows(min_row=2, min_col=7, max_col=9):
        for cell in row:
            cell.number_format = "0.0000"
    for row in metrics_ws.iter_rows(min_row=2, min_col=10, max_col=13):
        for cell in row:
            cell.number_format = "0.00"
    for row in metrics_ws.iter_rows(min_row=2, min_col=14, max_col=16):
        for cell in row:
            cell.number_format = "0.0000"

    if weekly_ws is not None:
        for row in weekly_ws.iter_rows(min_row=2, max_row=max(2, len(weekly_rows or []) + 1), min_col=3, max_col=9):
            for cell in row:
                if cell.column in (3, 5, 7, 9):
                    cell.number_format = "0.0000"
        for row in weekly_ws.iter_rows(min_row=2, max_row=max(2, len(weekly_rows or []) + 1), min_col=10, max_col=11):
            for cell in row:
                cell.number_format = "0.00"
        if benchmark_header_row:
            for row in weekly_ws.iter_rows(min_row=benchmark_header_row + 1, max_row=weekly_ws.max_row, min_col=3, max_col=8):
                for cell in row:
                    if cell.column in (3, 4, 5):
                        cell.number_format = "0.0000"
                    elif cell.column in (6, 8):
                        cell.number_format = "0.00"

    wb.save(filepath)


# ============== Tkinter GUI ==============
class FundAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("私募基金净值分析工具")
        self.root.geometry("1280x760")
        self.root.minsize(1120, 680)
        self.root.configure(bg="#f4f6f8")

        self.fund_name = ""
        self.filepath = ""
        self.data = []
        self.products = []
        self.selected_index = None
        self.rf_annual = RF_ANNUAL
        self.use_prev_year_end = True
        self.results = []
        self.one_year_data = None
        self.weekly_date_options = []
        self.weekly_date_vars = {}
        self.weekly_date_boxes = {}
        self.benchmark_vars = {}
        self.index_series_map = {}
        self.index_filepath = ""
        self.selected_index_name = tk.StringVar()
        self.chart_start_date_var = tk.StringVar()
        self.chart_end_date_var = tk.StringVar()
        self.chart_common_dates = []
        self.tree_sort_reverse = {}
        self.tree_heading_labels = {}
        self.dragged_weekly_item = None
        self.dragged_product_index = None
        self.dragged_heading = None
        self.next_product_id = 1

        self.setup_ui()

    def setup_ui(self):
        self.colors = {
            "bg": "#f4f6f8",
            "panel": "#ffffff",
            "navy": "#172033",
            "text": "#172033",
            "muted": "#667085",
            "line": "#d9dee8",
            "blue": "#0b3a75",
            "green": "#14532d",
            "red": "#b42318",
            "orange": "#7c2d12",
            "disabled": "#94a3b8",
        }

        try:
            style = ttk.Style()
            style.theme_use("clam")
            style.configure("Treeview", rowheight=28, font=("Arial", 11), borderwidth=0)
            style.configure("Treeview.Heading", font=("Arial", 11, "bold"), padding=(6, 8))
            style.map("Treeview", background=[("selected", "#dbeafe")], foreground=[("selected", self.colors["text"])])
        except tk.TclError:
            pass

        header_frame = tk.Frame(self.root, bg=self.colors["navy"], height=78)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)

        tk.Label(
            header_frame,
            text="私募基金净值分析工具",
            font=("Arial", 22, "bold"),
            bg=self.colors["navy"],
            fg="white",
        ).pack(side="left", padx=24)
        tk.Label(
            header_frame,
            text="上传净值 Excel，查看产品摘要、最新净值和区间业绩指标",
            font=("Arial", 12),
            bg=self.colors["navy"],
            fg="#d1d7e0",
        ).pack(side="left", padx=(8, 0), pady=(6, 0))

        main_frame = tk.Frame(self.root, bg=self.colors["bg"], padx=18, pady=18)
        main_frame.pack(fill="both", expand=True)

        left_frame = tk.Frame(main_frame, bg=self.colors["panel"], padx=16, pady=16, width=292)
        left_frame.pack(side="left", fill="y", padx=(0, 16))
        left_frame.pack_propagate(False)

        tk.Label(
            left_frame,
            text="文件与参数",
            font=("Arial", 16, "bold"),
            bg=self.colors["panel"],
            fg=self.colors["text"],
        ).pack(anchor="w")

        tk.Label(
            left_frame,
            text="支持 .xlsx/.xlsm；表头需包含净值日期、单位净值，可选累计净值或累计单位净值。",
            font=("Arial", 10),
            bg=self.colors["panel"],
            fg=self.colors["muted"],
            justify="left",
            wraplength=244,
        ).pack(anchor="w", pady=(6, 16))

        self.upload_btn = tk.Button(
            left_frame,
            text="批量上传净值 Excel 文件",
            command=self.upload_file,
            font=("Arial", 12, "bold"),
            bg=self.colors["blue"],
            fg="white",
            activebackground="#082f63",
            activeforeground="white",
            relief="flat",
            bd=0,
            highlightthickness=0,
            highlightbackground=self.colors["blue"],
            padx=14,
            pady=10,
            cursor="hand2",
        )
        self.upload_btn.pack(fill="x")

        self.file_label = tk.Label(
            left_frame,
            text="未选择文件",
            font=("Arial", 10),
            fg=self.colors["muted"],
            bg=self.colors["panel"],
            justify="left",
            wraplength=244,
        )
        self.file_label.pack(anchor="w", fill="x", pady=(10, 18))

        tk.Label(
            left_frame,
            text="产品列表",
            font=("Arial", 12, "bold"),
            bg=self.colors["panel"],
            fg=self.colors["text"],
        ).pack(anchor="w", pady=(0, 6))
        product_list_frame = tk.Frame(left_frame, bg=self.colors["panel"])
        product_list_frame.pack(fill="x", pady=(0, 16))
        self.product_listbox = tk.Listbox(
            product_list_frame,
            height=8,
            font=("Arial", 10),
            activestyle="none",
            relief="solid",
            bd=1,
            selectbackground="#dbeafe",
            selectforeground=self.colors["text"],
            exportselection=False,
        )
        self.product_listbox.pack(side="left", fill="both", expand=True)
        product_scrollbar = ttk.Scrollbar(product_list_frame, orient="vertical", command=self.product_listbox.yview)
        product_scrollbar.pack(side="right", fill="y")
        self.product_listbox.configure(yscrollcommand=product_scrollbar.set)
        self.product_listbox.bind("<<ListboxSelect>>", self.on_product_select)
        self.product_listbox.bind("<ButtonPress-1>", self.on_product_list_press)
        self.product_listbox.bind("<B1-Motion>", self.on_product_list_drag)
        self.product_listbox.bind("<ButtonRelease-1>", self.on_product_list_release)

        order_buttons = tk.Frame(left_frame, bg=self.colors["panel"])
        order_buttons.pack(fill="x", pady=(0, 16))
        tk.Button(
            order_buttons,
            text="上移",
            command=lambda: self.move_selected_product(-1),
            bg="#eef4ff",
            fg=self.colors["blue"],
            activebackground="#dbeafe",
            activeforeground=self.colors["blue"],
            relief="flat",
            bd=0,
            font=("Arial", 10, "bold"),
            cursor="hand2",
            padx=10,
            pady=6,
        ).pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Button(
            order_buttons,
            text="下移",
            command=lambda: self.move_selected_product(1),
            bg="#eef4ff",
            fg=self.colors["blue"],
            activebackground="#dbeafe",
            activeforeground=self.colors["blue"],
            relief="flat",
            bd=0,
            font=("Arial", 10, "bold"),
            cursor="hand2",
            padx=10,
            pady=6,
        ).pack(side="left", fill="x", expand=True)

        separator = tk.Frame(left_frame, bg=self.colors["line"], height=1)
        separator.pack(fill="x", pady=(0, 16))

        tk.Label(
            left_frame,
            text="计算设置",
            font=("Arial", 13, "bold"),
            bg=self.colors["panel"],
            fg=self.colors["text"],
        ).pack(anchor="w")

        tk.Label(
            left_frame,
            text="年化无风险利率 (%)",
            font=("Arial", 10),
            bg=self.colors["panel"],
            fg=self.colors["muted"],
        ).pack(anchor="w", pady=(12, 4))
        self.rf_entry = tk.Entry(left_frame, width=15, font=("Arial", 12), relief="solid", bd=1)
        self.rf_entry.insert(0, str(self.rf_annual))
        self.rf_entry.pack(anchor="w", fill="x", ipady=6)

        self.rf_var = tk.BooleanVar(value=True)
        self.rf_check = tk.Checkbutton(
            left_frame,
            text="年初基准使用前一年 12-31 净值",
            variable=self.rf_var,
            bg=self.colors["panel"],
            fg=self.colors["text"],
            activebackground=self.colors["panel"],
            font=("Arial", 10),
            anchor="w",
        )
        self.rf_check.pack(anchor="w", fill="x", pady=(10, 16))

        self.recalc_btn = tk.Button(
            left_frame,
            text="重新计算",
            command=self.recalculate,
            bg=self.colors["green"],
            fg="white",
            activebackground="#0f3f22",
            activeforeground="white",
            relief="flat",
            bd=0,
            highlightthickness=0,
            highlightbackground=self.colors["green"],
            font=("Arial", 11, "bold"),
            cursor="hand2",
            padx=14,
            pady=9,
        )
        self.recalc_btn.pack(fill="x", pady=(0, 8))

        self.export_btn = tk.Button(
            left_frame,
            text="导出 Excel 分析结果",
            command=self.export_results,
            bg=self.colors["orange"],
            fg="white",
            activebackground="#5f220d",
            activeforeground="white",
            disabledforeground="#f8fafc",
            relief="flat",
            bd=0,
            highlightthickness=0,
            highlightbackground=self.colors["orange"],
            font=("Arial", 11, "bold"),
            cursor="hand2",
            state="disabled",
            padx=14,
            pady=9,
        )
        self.export_btn.pack(fill="x")
        self.set_export_button_state(False)

        tk.Label(
            left_frame,
            text="提示：上传后右侧会直接显示计算结果；导出前会按当前参数重新计算。",
            font=("Arial", 10),
            bg=self.colors["panel"],
            fg=self.colors["muted"],
            justify="left",
            wraplength=244,
        ).pack(side="bottom", anchor="w")

        right_frame = tk.Frame(main_frame, bg=self.colors["bg"])
        right_frame.pack(side="left", fill="both", expand=True)

        self.notebook = ttk.Notebook(right_frame)
        self.notebook.pack(fill="both", expand=True)
        self.detail_tab = tk.Frame(self.notebook, bg=self.colors["bg"], padx=2, pady=2)
        self.compare_tab = tk.Frame(self.notebook, bg=self.colors["bg"], padx=2, pady=2)
        self.weekly_tab = tk.Frame(self.notebook, bg=self.colors["bg"], padx=2, pady=2)
        self.chart_tab = tk.Frame(self.notebook, bg=self.colors["bg"], padx=2, pady=2)
        self.notebook.add(self.detail_tab, text="单产品分析")
        self.notebook.add(self.compare_tab, text="产品对比")
        self.notebook.add(self.weekly_tab, text="周报涨幅")
        self.notebook.add(self.chart_tab, text="走势对比")

        product_frame = tk.Frame(self.detail_tab, bg=self.colors["panel"], padx=18, pady=16)
        product_frame.pack(fill="x", pady=(0, 14))

        self.product_name_label = tk.Label(
            product_frame,
            text="请上传净值 Excel 文件",
            font=("Arial", 18, "bold"),
            bg=self.colors["panel"],
            fg=self.colors["text"],
            anchor="w",
            justify="left",
            wraplength=820,
        )
        self.product_name_label.pack(anchor="w", fill="x")

        self.product_meta_label = tk.Label(
            product_frame,
            text="上传后会在这里显示产品名称、数据区间、记录数和计算参数。",
            font=("Arial", 11),
            bg=self.colors["panel"],
            fg=self.colors["muted"],
            anchor="w",
            justify="left",
            wraplength=920,
        )
        self.product_meta_label.pack(anchor="w", fill="x", pady=(6, 12))

        summary_grid = tk.Frame(product_frame, bg=self.colors["panel"])
        summary_grid.pack(fill="x")
        self.summary_labels = {}
        summary_items = [
            ("file", "文件"),
            ("period", "数据区间"),
            ("count", "记录数"),
            ("rf", "无风险利率"),
        ]
        for col, (key, title) in enumerate(summary_items):
            cell = tk.Frame(summary_grid, bg=self.colors["panel"])
            cell.grid(row=0, column=col, sticky="ew", padx=(0 if col == 0 else 16, 0))
            summary_grid.columnconfigure(col, weight=1)
            tk.Label(cell, text=title, font=("Arial", 10), bg=self.colors["panel"], fg=self.colors["muted"]).pack(anchor="w")
            self.summary_labels[key] = tk.Label(
                cell,
                text="--",
                font=("Arial", 11, "bold"),
                bg=self.colors["panel"],
                fg=self.colors["text"],
                anchor="w",
                justify="left",
                wraplength=200,
            )
            self.summary_labels[key].pack(anchor="w")

        metric_frame = tk.Frame(self.detail_tab, bg=self.colors["bg"])
        metric_frame.pack(fill="x", pady=(0, 14))
        self.metric_value_labels = {}
        self.metric_hint_labels = {}
        metric_items = [
            ("latest_nav", "最新单位净值"),
            ("cum_nav", "最新累计净值"),
            ("one_year", "近一年收益率"),
            ("latest_date", "最新净值日期"),
        ]
        for col, (key, title) in enumerate(metric_items):
            card = tk.Frame(metric_frame, bg=self.colors["panel"], padx=16, pady=14)
            card.grid(row=0, column=col, sticky="nsew", padx=(0 if col == 0 else 12, 0))
            metric_frame.columnconfigure(col, weight=1, uniform="metric")
            tk.Label(card, text=title, font=("Arial", 11), bg=self.colors["panel"], fg=self.colors["muted"]).pack(anchor="w")
            self.metric_value_labels[key] = tk.Label(
                card,
                text="--",
                font=("Arial", 22, "bold"),
                bg=self.colors["panel"],
                fg=self.colors["text"],
                anchor="w",
            )
            self.metric_value_labels[key].pack(anchor="w", pady=(6, 2))
            self.metric_hint_labels[key] = tk.Label(
                card,
                text="上传后显示",
                font=("Arial", 10),
                bg=self.colors["panel"],
                fg=self.colors["muted"],
                anchor="w",
            )
            self.metric_hint_labels[key].pack(anchor="w")

        self.metrics_frame = tk.Frame(self.detail_tab, bg=self.colors["panel"], padx=14, pady=12)
        self.metrics_frame.pack(fill="both", expand=True)
        table_header = tk.Frame(self.metrics_frame, bg=self.colors["panel"])
        table_header.pack(fill="x", pady=(0, 8))
        tk.Label(
            table_header,
            text="区间业绩指标",
            font=("Arial", 15, "bold"),
            bg=self.colors["panel"],
            fg=self.colors["text"],
        ).pack(side="left")
        self.table_note_label = tk.Label(
            table_header,
            text="上传后显示逐年和成立以来指标",
            font=("Arial", 10),
            bg=self.colors["panel"],
            fg=self.colors["muted"],
        )
        self.table_note_label.pack(side="left", padx=(12, 0), pady=(4, 0))

        table_container = tk.Frame(self.metrics_frame, bg=self.colors["panel"])
        table_container.pack(fill="both", expand=True)

        self.tree_columns = [
            ("label", "区间", 120),
            ("start_date", "起始日期", 92),
            ("end_date", "结束日期", 92),
            ("start_nav", "期初单位净值", 112),
            ("end_nav", "期末单位净值", 112),
            ("cum_nav", "期末累计净值", 112),
            ("total_return", "区间收益%", 92),
            ("ann_return", "年化收益%", 92),
            ("max_drawdown", "最大回撤%", 92),
            ("ann_volatility", "年化波动率%", 100),
            ("sharpe_ratio", "Sharpe", 80),
            ("sortino_ratio", "Sortino", 80),
            ("calmar_ratio", "Calmar", 80),
            ("note", "备注", 170),
        ]
        columns = [col_id for col_id, _, _ in self.tree_columns]
        self.tree = ttk.Treeview(table_container, columns=columns, show="headings", height=14)

        for col_id, heading, width in self.tree_columns:
            self.tree.heading(col_id, text=heading)
            self.tree.column(col_id, width=width, minwidth=width, anchor="center", stretch=False)
        self.enable_tree_sorting(self.tree, self.tree_columns)

        y_scrollbar = ttk.Scrollbar(table_container, orient="vertical", command=self.tree.yview)
        x_scrollbar = ttk.Scrollbar(table_container, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
        self.tree.tag_configure("positive", foreground=self.colors["red"])
        self.tree.tag_configure("negative", foreground=self.colors["green"])

        self.tree.grid(row=0, column=0, sticky="nsew")
        y_scrollbar.grid(row=0, column=1, sticky="ns")
        x_scrollbar.grid(row=1, column=0, sticky="ew")
        table_container.rowconfigure(0, weight=1)
        table_container.columnconfigure(0, weight=1)

        compare_frame = tk.Frame(self.compare_tab, bg=self.colors["panel"], padx=14, pady=12)
        compare_frame.pack(fill="both", expand=True)
        compare_header = tk.Frame(compare_frame, bg=self.colors["panel"])
        compare_header.pack(fill="x", pady=(0, 8))
        tk.Label(
            compare_header,
            text="产品对比",
            font=("Arial", 15, "bold"),
            bg=self.colors["panel"],
            fg=self.colors["text"],
        ).pack(side="left")
        self.compare_note_label = tk.Label(
            compare_header,
            text="批量上传后显示所有产品核心指标",
            font=("Arial", 10),
            bg=self.colors["panel"],
            fg=self.colors["muted"],
        )
        self.compare_note_label.pack(side="left", padx=(12, 0), pady=(4, 0))

        compare_table_container = tk.Frame(compare_frame, bg=self.colors["panel"])
        compare_table_container.pack(fill="both", expand=True)
        self.compare_columns = [
            ("fund", "产品", 260),
            ("file", "文件", 150),
            ("start_date", "数据起点", 92),
            ("latest_date", "最新日期", 92),
            ("records", "记录数", 72),
            ("latest_nav", "单位净值", 82),
            ("latest_cum_nav", "累计净值", 82),
            ("near_one_return", "近一年%", 82),
            ("near_two_return", "近2年%", 82),
            ("near_three_return", "近3年%", 82),
            ("since_return", "成立以来%", 92),
            ("since_max_drawdown", "最大回撤%", 92),
            ("since_sharpe", "Sharpe", 78),
            ("note", "备注", 240),
        ]
        compare_ids = [col_id for col_id, _, _ in self.compare_columns]
        self.compare_tree = ttk.Treeview(compare_table_container, columns=compare_ids, show="headings", height=18)
        for col_id, heading, width in self.compare_columns:
            self.compare_tree.heading(col_id, text=heading)
            self.compare_tree.column(col_id, width=width, minwidth=width, anchor="center", stretch=False)
        self.enable_tree_sorting(self.compare_tree, self.compare_columns)
        compare_y = ttk.Scrollbar(compare_table_container, orient="vertical", command=self.compare_tree.yview)
        compare_x = ttk.Scrollbar(compare_table_container, orient="horizontal", command=self.compare_tree.xview)
        self.compare_tree.configure(yscrollcommand=compare_y.set, xscrollcommand=compare_x.set)
        self.compare_tree.grid(row=0, column=0, sticky="nsew")
        compare_y.grid(row=0, column=1, sticky="ns")
        compare_x.grid(row=1, column=0, sticky="ew")
        compare_table_container.rowconfigure(0, weight=1)
        compare_table_container.columnconfigure(0, weight=1)

        self.setup_weekly_tab()
        self.setup_chart_tab()

    def setup_weekly_tab(self):
        control_panel = tk.Frame(self.weekly_tab, bg=self.colors["panel"], padx=14, pady=12)
        control_panel.pack(fill="x", pady=(0, 14))

        header = tk.Frame(control_panel, bg=self.colors["panel"])
        header.pack(fill="x", pady=(0, 10))
        tk.Label(
            header,
            text="周报涨幅",
            font=("Arial", 15, "bold"),
            bg=self.colors["panel"],
            fg=self.colors["text"],
        ).pack(side="left")
        self.weekly_note_label = tk.Label(
            header,
            text="上传后选择周报日期，并填写中证1000 / 中证800收盘价",
            font=("Arial", 10),
            bg=self.colors["panel"],
            fg=self.colors["muted"],
        )
        self.weekly_note_label.pack(side="left", padx=(12, 0), pady=(4, 0))

        date_frame = tk.Frame(control_panel, bg=self.colors["panel"])
        date_frame.pack(fill="x", pady=(0, 10))
        date_fields = [
            ("unit", "单位净值日期 / 区间起点"),
            ("prev", "上周收盘日期"),
            ("current", "本周收盘日期 / 区间终点"),
            ("cum", "累计净值日期"),
        ]
        for col, (key, label) in enumerate(date_fields):
            cell = tk.Frame(date_frame, bg=self.colors["panel"])
            cell.grid(row=0, column=col, sticky="ew", padx=(0 if col == 0 else 12, 0))
            date_frame.columnconfigure(col, weight=1, uniform="weekly_dates")
            tk.Label(
                cell,
                text=label,
                font=("Arial", 10),
                bg=self.colors["panel"],
                fg=self.colors["muted"],
            ).pack(anchor="w", pady=(0, 4))
            var = tk.StringVar()
            combo = ttk.Combobox(cell, textvariable=var, state="readonly", font=("Arial", 11), width=14)
            combo.pack(fill="x", ipady=2)
            combo.bind("<<ComboboxSelected>>", self.update_weekly_display)
            self.weekly_date_vars[key] = var
            self.weekly_date_boxes[key] = combo

        benchmark_frame = tk.Frame(control_panel, bg=self.colors["panel"])
        benchmark_frame.pack(fill="x")

        input_frame = tk.Frame(benchmark_frame, bg=self.colors["panel"])
        input_frame.pack(side="left", fill="x", expand=True, padx=(0, 14))
        tk.Label(
            input_frame,
            text="对标指数数据",
            font=("Arial", 12, "bold"),
            bg=self.colors["panel"],
            fg=self.colors["text"],
        ).grid(row=0, column=0, columnspan=5, sticky="w", pady=(0, 6))
        tk.Label(
            input_frame,
            text="填写收盘价，程序自动计算本周涨幅和区间涨幅",
            font=("Arial", 10),
            bg=self.colors["panel"],
            fg=self.colors["muted"],
        ).grid(row=0, column=2, columnspan=5, sticky="e", pady=(0, 6))

        for row_idx, name in enumerate(BENCHMARK_NAMES, start=1):
            tk.Label(
                input_frame,
                text=name,
                font=("Arial", 11, "bold"),
                bg=self.colors["panel"],
                fg=self.colors["text"],
                width=9,
                anchor="w",
            ).grid(row=row_idx, column=0, sticky="w", pady=3)

            start_var = tk.StringVar()
            prev_var = tk.StringVar()
            current_var = tk.StringVar()
            self.benchmark_vars[name] = {"start": start_var, "prev": prev_var, "current": current_var}

            tk.Label(
                input_frame,
                text="区间起点收盘",
                font=("Arial", 10),
                bg=self.colors["panel"],
                fg=self.colors["muted"],
            ).grid(row=row_idx, column=1, sticky="e", padx=(8, 4))
            start_entry = tk.Entry(input_frame, textvariable=start_var, font=("Arial", 11), width=10, relief="solid", bd=1)
            start_entry.grid(row=row_idx, column=2, sticky="w", ipady=4)
            start_entry.bind("<KeyRelease>", self.update_weekly_display)

            tk.Label(
                input_frame,
                text="上周收盘",
                font=("Arial", 10),
                bg=self.colors["panel"],
                fg=self.colors["muted"],
            ).grid(row=row_idx, column=3, sticky="e", padx=(14, 4))
            prev_entry = tk.Entry(input_frame, textvariable=prev_var, font=("Arial", 11), width=10, relief="solid", bd=1)
            prev_entry.grid(row=row_idx, column=4, sticky="w", ipady=4)
            prev_entry.bind("<KeyRelease>", self.update_weekly_display)

            tk.Label(
                input_frame,
                text="本周收盘",
                font=("Arial", 10),
                bg=self.colors["panel"],
                fg=self.colors["muted"],
            ).grid(row=row_idx, column=5, sticky="e", padx=(14, 4))
            current_entry = tk.Entry(input_frame, textvariable=current_var, font=("Arial", 11), width=10, relief="solid", bd=1)
            current_entry.grid(row=row_idx, column=6, sticky="w", ipady=4)
            current_entry.bind("<KeyRelease>", self.update_weekly_display)

        refresh_btn = tk.Button(
            benchmark_frame,
            text="刷新周报",
            command=self.update_weekly_display,
            bg=self.colors["green"],
            fg="white",
            activebackground="#0f3f22",
            activeforeground="white",
            relief="flat",
            bd=0,
            highlightthickness=0,
            font=("Arial", 11, "bold"),
            cursor="hand2",
            padx=16,
            pady=9,
        )
        refresh_btn.pack(side="right", anchor="n")

        benchmark_display = tk.Frame(control_panel, bg=self.colors["panel"])
        benchmark_display.pack(fill="x", pady=(10, 0))
        benchmark_columns = [
            ("name", "指数", 100),
            ("start_close", "区间起点收盘", 120),
            ("prev_close", "上周收盘", 100),
            ("current_close", "本周收盘", 100),
            ("weekly_return", "本周涨幅%", 92),
            ("interval_return", "区间涨幅%", 92),
        ]
        benchmark_ids = [col_id for col_id, _, _ in benchmark_columns]
        self.benchmark_tree = ttk.Treeview(benchmark_display, columns=benchmark_ids, show="headings", height=2)
        for col_id, heading, width in benchmark_columns:
            self.benchmark_tree.heading(col_id, text=heading)
            self.benchmark_tree.column(col_id, width=width, minwidth=width, anchor="center", stretch=False)
        self.enable_tree_sorting(self.benchmark_tree, benchmark_columns)
        self.benchmark_tree.pack(fill="x")

        weekly_table_frame = tk.Frame(self.weekly_tab, bg=self.colors["panel"], padx=14, pady=12)
        weekly_table_frame.pack(fill="both", expand=True)
        table_header = tk.Frame(weekly_table_frame, bg=self.colors["panel"])
        table_header.pack(fill="x", pady=(0, 8))
        tk.Label(
            table_header,
            text="全部产品周报明细",
            font=("Arial", 15, "bold"),
            bg=self.colors["panel"],
            fg=self.colors["text"],
        ).pack(side="left")

        table_container = tk.Frame(weekly_table_frame, bg=self.colors["panel"])
        table_container.pack(fill="both", expand=True)
        self.weekly_columns = [
            ("fund", "产品", 260),
            ("unit_date", "单位净值日期", 100),
            ("unit_nav", "单位净值", 90),
            ("prev_date", "上周收盘日期", 100),
            ("prev_nav", "上周收盘单位净值", 138),
            ("current_date", "本周收盘日期", 100),
            ("current_nav", "本周收盘单位净值", 138),
            ("cum_date", "累计净值日期", 100),
            ("cum_nav", "累计净值", 90),
            ("weekly_return", "本周涨幅%", 90),
            ("interval_return", "区间收益率%", 100),
            ("benchmark_result", "是否跑赢对标指数", 310),
            ("note", "备注", 260),
        ]
        weekly_ids = [col_id for col_id, _, _ in self.weekly_columns]
        self.weekly_tree = ttk.Treeview(table_container, columns=weekly_ids, show="headings", height=14)
        for col_id, heading, width in self.weekly_columns:
            self.weekly_tree.heading(col_id, text=heading)
            self.weekly_tree.column(col_id, width=width, minwidth=width, anchor="center", stretch=False)
        self.enable_tree_sorting(self.weekly_tree, self.weekly_columns)
        weekly_y = ttk.Scrollbar(table_container, orient="vertical", command=self.weekly_tree.yview)
        weekly_x = ttk.Scrollbar(table_container, orient="horizontal", command=self.weekly_tree.xview)
        self.weekly_tree.configure(yscrollcommand=weekly_y.set, xscrollcommand=weekly_x.set)
        self.weekly_tree.tag_configure("positive", foreground=self.colors["red"])
        self.weekly_tree.tag_configure("negative", foreground=self.colors["green"])
        self.weekly_tree.bind("<ButtonPress-1>", self.on_weekly_tree_press, add="+")
        self.weekly_tree.bind("<B1-Motion>", self.on_weekly_tree_drag, add="+")
        self.weekly_tree.grid(row=0, column=0, sticky="nsew")
        weekly_y.grid(row=0, column=1, sticky="ns")
        weekly_x.grid(row=1, column=0, sticky="ew")
        table_container.rowconfigure(0, weight=1)
        table_container.columnconfigure(0, weight=1)

    def setup_chart_tab(self):
        control_panel = tk.Frame(self.chart_tab, bg=self.colors["panel"], padx=14, pady=12)
        control_panel.pack(fill="x", pady=(0, 14))

        header = tk.Frame(control_panel, bg=self.colors["panel"])
        header.pack(fill="x", pady=(0, 10))
        tk.Label(
            header,
            text="净值走势对比",
            font=("Arial", 15, "bold"),
            bg=self.colors["panel"],
            fg=self.colors["text"],
        ).pack(side="left")
        self.chart_note_label = tk.Label(
            header,
            text="上传指数 Excel 后，当前产品会和指数换算值画在同一张图里",
            font=("Arial", 10),
            bg=self.colors["panel"],
            fg=self.colors["muted"],
        )
        self.chart_note_label.pack(side="left", padx=(12, 0), pady=(4, 0))

        chart_controls = tk.Frame(control_panel, bg=self.colors["panel"])
        chart_controls.pack(fill="x")
        tk.Button(
            chart_controls,
            text="上传指数 Excel",
            command=self.upload_index_file,
            bg=self.colors["blue"],
            fg="white",
            activebackground="#082f63",
            activeforeground="white",
            relief="flat",
            bd=0,
            highlightthickness=0,
            font=("Arial", 11, "bold"),
            cursor="hand2",
            padx=14,
            pady=8,
        ).pack(side="left")

        self.index_file_label = tk.Label(
            chart_controls,
            text="未上传指数文件",
            font=("Arial", 10),
            bg=self.colors["panel"],
            fg=self.colors["muted"],
            anchor="w",
        )
        self.index_file_label.pack(side="left", padx=(12, 18))

        tk.Label(
            chart_controls,
            text="选择指数",
            font=("Arial", 10),
            bg=self.colors["panel"],
            fg=self.colors["muted"],
        ).pack(side="left", padx=(0, 6))
        self.index_combo = ttk.Combobox(
            chart_controls,
            textvariable=self.selected_index_name,
            state="readonly",
            font=("Arial", 11),
            width=24,
        )
        self.index_combo.pack(side="left", ipady=2)
        self.index_combo.bind("<<ComboboxSelected>>", self.on_chart_index_select)

        tk.Label(
            chart_controls,
            text="开始",
            font=("Arial", 10),
            bg=self.colors["panel"],
            fg=self.colors["muted"],
        ).pack(side="left", padx=(14, 6))
        self.chart_start_combo = ttk.Combobox(
            chart_controls,
            textvariable=self.chart_start_date_var,
            state="readonly",
            font=("Arial", 11),
            width=12,
        )
        self.chart_start_combo.pack(side="left", ipady=2)
        self.chart_start_combo.bind("<<ComboboxSelected>>", self.update_chart_display)

        tk.Label(
            chart_controls,
            text="结束",
            font=("Arial", 10),
            bg=self.colors["panel"],
            fg=self.colors["muted"],
        ).pack(side="left", padx=(10, 6))
        self.chart_end_combo = ttk.Combobox(
            chart_controls,
            textvariable=self.chart_end_date_var,
            state="readonly",
            font=("Arial", 11),
            width=12,
        )
        self.chart_end_combo.pack(side="left", ipady=2)
        self.chart_end_combo.bind("<<ComboboxSelected>>", self.update_chart_display)

        charts_frame = tk.Frame(self.chart_tab, bg=self.colors["bg"])
        charts_frame.pack(fill="both", expand=True)

        unit_panel = tk.Frame(charts_frame, bg=self.colors["panel"], padx=12, pady=10)
        unit_panel.pack(fill="both", expand=True, pady=(0, 12))
        tk.Label(
            unit_panel,
            text="单位净值 vs 指数换算值",
            font=("Arial", 13, "bold"),
            bg=self.colors["panel"],
            fg=self.colors["text"],
        ).pack(anchor="w", pady=(0, 6))
        self.unit_chart_canvas = tk.Canvas(unit_panel, bg="white", height=230, highlightthickness=1, highlightbackground=self.colors["line"])
        self.unit_chart_canvas.pack(fill="both", expand=True)
        self.unit_chart_canvas.bind("<Configure>", lambda _event: self.update_chart_display())

        cum_panel = tk.Frame(charts_frame, bg=self.colors["panel"], padx=12, pady=10)
        cum_panel.pack(fill="both", expand=True)
        tk.Label(
            cum_panel,
            text="累计净值 vs 指数换算值",
            font=("Arial", 13, "bold"),
            bg=self.colors["panel"],
            fg=self.colors["text"],
        ).pack(anchor="w", pady=(0, 6))
        self.cum_chart_canvas = tk.Canvas(cum_panel, bg="white", height=230, highlightthickness=1, highlightbackground=self.colors["line"])
        self.cum_chart_canvas.pack(fill="both", expand=True)
        self.cum_chart_canvas.bind("<Configure>", lambda _event: self.update_chart_display())

    def enable_tree_sorting(self, tree, column_defs):
        tree_key = str(tree)
        self.tree_heading_labels[tree_key] = {col_id: heading for col_id, heading, _ in column_defs}
        for col_id, heading, _ in column_defs:
            tree.heading(col_id, text=f"{heading} ↕")
        tree.bind("<ButtonPress-1>", lambda event, t=tree: self.on_tree_button_press(event, t), add="+")
        tree.bind("<B1-Motion>", lambda event, t=tree: self.on_tree_heading_drag(event, t), add="+")
        tree.bind("<ButtonRelease-1>", lambda event, t=tree: self.on_tree_button_release(event, t), add="+")

    def get_display_columns(self, tree):
        display_columns = tree["displaycolumns"]
        if display_columns in ("#all", ("#all",), ""):
            return list(tree["columns"])
        if isinstance(display_columns, str):
            return display_columns.split()
        return list(display_columns)

    def get_tree_column_from_x(self, tree, x):
        column_ref = tree.identify_column(x)
        if not column_ref:
            return None
        try:
            display_index = int(column_ref.replace("#", "")) - 1
        except ValueError:
            return None
        display_columns = self.get_display_columns(tree)
        if 0 <= display_index < len(display_columns):
            return display_columns[display_index]
        return None

    def reorder_tree_column(self, tree, source_column, target_column):
        if not source_column or not target_column or source_column == target_column:
            return False
        display_columns = self.get_display_columns(tree)
        if source_column not in display_columns or target_column not in display_columns:
            return False
        source_index = display_columns.index(source_column)
        target_index = display_columns.index(target_column)
        display_columns.remove(source_column)
        target_index = display_columns.index(target_column)
        if source_index < target_index:
            target_index += 1
        display_columns.insert(target_index, source_column)
        tree["displaycolumns"] = display_columns
        return True

    def get_sort_value(self, value):
        text = str(value or "").strip()
        if text in ("", "--", "无", "数据不足"):
            return (1, "")

        parsed_date = parse_date(text)
        if parsed_date:
            return (0, parsed_date.toordinal())

        cleaned = re.sub(r"(pct|%|％)", "", text, flags=re.IGNORECASE)
        cleaned = cleaned.replace(",", "").replace("+", "").strip()
        try:
            return (0, float(cleaned))
        except ValueError:
            return (0, text)

    def sort_treeview(self, tree, column):
        sort_key = (str(tree), column)
        reverse = self.tree_sort_reverse.get(sort_key, False)
        rows = [(self.get_sort_value(tree.set(item, column)), item) for item in tree.get_children("")]
        rows.sort(key=lambda item: item[0], reverse=reverse)
        for index, (_, item) in enumerate(rows):
            tree.move(item, "", index)
        self.tree_sort_reverse[sort_key] = not reverse
        self.update_tree_sort_headers(tree, column, "↓" if reverse else "↑")

        if tree in (getattr(self, "weekly_tree", None), getattr(self, "compare_tree", None)):
            self.sync_products_from_tree_order(tree)

    def update_tree_sort_headers(self, tree, active_column, direction):
        tree_key = str(tree)
        for col_id, heading in self.tree_heading_labels.get(tree_key, {}).items():
            suffix = direction if col_id == active_column else "↕"
            tree.heading(col_id, text=f"{heading} {suffix}")

    def on_tree_button_press(self, event, tree):
        if tree.identify_region(event.x, event.y) != "heading":
            return None
        column = self.get_tree_column_from_x(tree, event.x)
        if not column:
            return "break"
        self.dragged_heading = {
            "tree": tree,
            "column": column,
            "start_x": event.x,
            "dragged": False,
        }
        return "break"

    def on_tree_heading_drag(self, event, tree):
        if not self.dragged_heading or self.dragged_heading.get("tree") is not tree:
            return None
        if abs(event.x - self.dragged_heading["start_x"]) < 8:
            return "break"

        target_column = self.get_tree_column_from_x(tree, event.x)
        if target_column and self.reorder_tree_column(tree, self.dragged_heading["column"], target_column):
            self.dragged_heading["dragged"] = True
        return "break"

    def on_tree_button_release(self, event, tree):
        if self.dragged_heading and self.dragged_heading.get("tree") is tree:
            column = self.dragged_heading["column"]
            dragged = self.dragged_heading.get("dragged", False)
            self.dragged_heading = None
            if dragged:
                return "break"
            if tree.identify_region(event.x, event.y) == "heading":
                self.sort_treeview(tree, column)
                return "break"

        region = tree.identify_region(event.x, event.y)
        if region == "heading":
            column = self.get_tree_column_from_x(tree, event.x)
            if column:
                self.sort_treeview(tree, column)
            return "break"

        if tree is getattr(self, "weekly_tree", None) and self.dragged_weekly_item:
            self.sync_products_from_tree_order(tree)
            self.dragged_weekly_item = None
            return "break"
        return None

    def sync_products_from_tree_order(self, tree):
        ordered_ids = [item for item in tree.get_children("") if str(item).startswith("p")]
        if not ordered_ids:
            return

        selected_product = self.get_selected_product()
        selected_id = selected_product.get("product_id") if selected_product else None
        product_by_id = {product.get("product_id"): product for product in self.products}
        ordered_products = [product_by_id[product_id] for product_id in ordered_ids if product_id in product_by_id]
        remaining = [product for product in self.products if product.get("product_id") not in ordered_ids]
        if ordered_products:
            self.products = ordered_products + remaining
            if selected_id:
                for idx, product in enumerate(self.products):
                    if product.get("product_id") == selected_id:
                        self.selected_index = idx
                        break
            self.update_product_list()

    def on_weekly_tree_press(self, event):
        if self.weekly_tree.identify_region(event.x, event.y) == "heading":
            self.dragged_weekly_item = None
            return
        item = self.weekly_tree.identify_row(event.y)
        self.dragged_weekly_item = item or None
        if item:
            self.weekly_tree.selection_set(item)

    def on_weekly_tree_drag(self, event):
        if not self.dragged_weekly_item:
            return
        target = self.weekly_tree.identify_row(event.y)
        if not target or target == self.dragged_weekly_item:
            return
        target_index = self.weekly_tree.index(target)
        self.weekly_tree.move(self.dragged_weekly_item, "", target_index)
        self.weekly_tree.selection_set(self.dragged_weekly_item)

    def upload_index_file(self):
        filepath = filedialog.askopenfilename(
            title="选择指数 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")]
        )
        if not filepath:
            return

        try:
            self.index_series_map = load_index_data(filepath)
        except Exception as exc:
            messagebox.showerror("错误", f"指数文件读取失败：\n{exc}")
            return

        self.index_filepath = filepath
        names = sorted(self.index_series_map.keys())
        self.index_combo.configure(values=names)
        if names:
            self.selected_index_name.set(names[0])
        self.index_file_label.config(text=f"{Path(filepath).name}，共 {len(names)} 个指数")
        self.update_chart_date_options(reset=True)
        self.update_chart_display()
        self.notebook.select(self.chart_tab)

    def get_chart_common_dates(self, product=None, index_data=None):
        product = product or self.get_selected_product()
        if index_data is None:
            index_data = self.index_series_map.get(self.selected_index_name.get())
        if not product or not index_data:
            return []

        product_dates = {record[0] for record in product["data"]}
        index_dates = {record[0] for record in index_data}
        return sorted(product_dates & index_dates)

    def update_chart_date_options(self, reset=False):
        if not hasattr(self, "chart_start_combo"):
            return

        product = self.get_selected_product()
        index_data = self.index_series_map.get(self.selected_index_name.get())
        common_dates = self.get_chart_common_dates(product, index_data)
        self.chart_common_dates = common_dates
        values = [str(d) for d in common_dates]
        self.chart_start_combo.configure(values=values)
        self.chart_end_combo.configure(values=values)

        if not values:
            self.chart_start_date_var.set("")
            self.chart_end_date_var.set("")
            return

        start_current = self.chart_start_date_var.get()
        end_current = self.chart_end_date_var.get()
        if reset or start_current not in values:
            self.chart_start_date_var.set(values[0])
        if reset or end_current not in values:
            self.chart_end_date_var.set(values[-1])

    def get_chart_date_range(self):
        start_date = parse_date(self.chart_start_date_var.get())
        end_date = parse_date(self.chart_end_date_var.get())
        if not start_date or not end_date:
            return None, None
        if start_date > end_date:
            return end_date, start_date
        return start_date, end_date

    def on_chart_index_select(self, _event=None):
        self.update_chart_date_options(reset=True)
        self.update_chart_display()

    def build_chart_series(self, product, index_data, value_type, start_date=None, end_date=None):
        if not product or not index_data:
            return [], [], ""

        index_by_date = {d: value for d, value in index_data}
        product_points = []
        for record in product["data"]:
            d = record[0]
            if start_date and d < start_date:
                continue
            if end_date and d > end_date:
                continue
            if d not in index_by_date:
                continue
            value = record[1] if value_type == "unit" else (record[2] if len(record) > 2 else None)
            if value is not None and value > 0:
                product_points.append((d, value))

        if len(product_points) < 2:
            return [], [], "产品累计净值数据不足" if value_type == "cum" else "产品单位净值数据不足"

        chart_points = []
        index_points = []
        base_product = None
        base_index = None
        base_date = None

        for d, product_value in product_points:
            index_value = index_by_date.get(d)
            if index_value is None:
                continue
            if base_product is None:
                base_product = product_value
                base_index = index_value
                base_date = d
            if base_index and base_index > 0:
                chart_points.append((d, product_value))
                index_points.append((d, index_value / base_index * base_product))

        if len(chart_points) < 2:
            return [], [], "所选日期范围内共同日期不足"

        range_text = f"{start_date} 至 {end_date}；" if start_date and end_date else ""
        note = f"{range_text}指数以 {base_date} 为起点换算，起点值对齐产品净值 {base_product:.4f}"
        return chart_points, index_points, note

    def draw_line_chart(self, canvas, title, product_points, index_points, product_label, index_label, empty_text="暂无可绘制数据"):
        canvas.delete("all")
        width = max(canvas.winfo_width(), 680)
        height = max(canvas.winfo_height(), 220)

        if len(product_points) < 2 or len(index_points) < 2:
            canvas.create_text(
                width / 2,
                height / 2,
                text=empty_text,
                fill=self.colors["muted"],
                font=("Arial", 12),
            )
            return

        margin_left = 62
        margin_right = 28
        margin_top = 28
        margin_bottom = 38
        plot_w = width - margin_left - margin_right
        plot_h = height - margin_top - margin_bottom

        all_points = product_points + index_points
        min_date = min(d for d, _ in all_points)
        max_date = max(d for d, _ in all_points)
        min_val = min(v for _, v in all_points)
        max_val = max(v for _, v in all_points)
        if min_val == max_val:
            min_val *= 0.98
            max_val *= 1.02
        padding = (max_val - min_val) * 0.08
        min_val -= padding
        max_val += padding
        date_span = max((max_date - min_date).days, 1)

        def to_xy(point):
            d, value = point
            x = margin_left + ((d - min_date).days / date_span) * plot_w
            y = margin_top + (1 - (value - min_val) / (max_val - min_val)) * plot_h
            return x, y

        canvas.create_text(margin_left, 12, text=title, anchor="w", fill=self.colors["text"], font=("Arial", 12, "bold"))
        for i in range(5):
            y = margin_top + plot_h * i / 4
            value = max_val - (max_val - min_val) * i / 4
            canvas.create_line(margin_left, y, width - margin_right, y, fill="#eef2f7")
            canvas.create_text(margin_left - 8, y, text=f"{value:.2f}", anchor="e", fill=self.colors["muted"], font=("Arial", 9))

        canvas.create_line(margin_left, margin_top, margin_left, height - margin_bottom, fill=self.colors["line"])
        canvas.create_line(margin_left, height - margin_bottom, width - margin_right, height - margin_bottom, fill=self.colors["line"])
        canvas.create_text(margin_left, height - 16, text=str(min_date), anchor="w", fill=self.colors["muted"], font=("Arial", 9))
        canvas.create_text(width - margin_right, height - 16, text=str(max_date), anchor="e", fill=self.colors["muted"], font=("Arial", 9))

        def draw_series(points, color):
            coords = []
            for point in points:
                coords.extend(to_xy(point))
            if len(coords) >= 4:
                canvas.create_line(*coords, fill=color, width=2.4, smooth=True)
            for point in (points[0], points[-1]):
                x, y = to_xy(point)
                canvas.create_oval(x - 3, y - 3, x + 3, y + 3, fill=color, outline=color)

        product_color = self.colors["blue"]
        index_color = self.colors["orange"]
        draw_series(product_points, product_color)
        draw_series(index_points, index_color)

        legend_x = width - margin_right - 250
        canvas.create_line(legend_x, 16, legend_x + 24, 16, fill=product_color, width=3)
        canvas.create_text(legend_x + 30, 16, text=product_label, anchor="w", fill=self.colors["text"], font=("Arial", 9))
        canvas.create_line(legend_x + 132, 16, legend_x + 156, 16, fill=index_color, width=3)
        canvas.create_text(legend_x + 162, 16, text=index_label, anchor="w", fill=self.colors["text"], font=("Arial", 9))

    def update_chart_display(self, _event=None):
        if not hasattr(self, "unit_chart_canvas"):
            return

        product = self.get_selected_product()
        index_name = self.selected_index_name.get()
        index_data = self.index_series_map.get(index_name)

        if not product:
            message = "请先上传产品净值 Excel"
            self.draw_line_chart(self.unit_chart_canvas, "单位净值走势", [], [], "", "", message)
            self.draw_line_chart(self.cum_chart_canvas, "累计净值走势", [], [], "", "", message)
            self.chart_note_label.config(text=message)
            return

        if not index_data:
            message = "请在本页上传指数 Excel 并选择指数"
            self.draw_line_chart(self.unit_chart_canvas, "单位净值走势", [], [], "", "", message)
            self.draw_line_chart(self.cum_chart_canvas, "累计净值走势", [], [], "", "", message)
            self.chart_note_label.config(text=message)
            return

        self.update_chart_date_options()
        start_date, end_date = self.get_chart_date_range()
        if not start_date or not end_date:
            message = "当前产品和指数没有共同日期，无法选择日期范围"
            self.draw_line_chart(self.unit_chart_canvas, "单位净值走势", [], [], "", "", message)
            self.draw_line_chart(self.cum_chart_canvas, "累计净值走势", [], [], "", "", message)
            self.chart_note_label.config(text=message)
            return

        unit_points, unit_index_points, unit_note = self.build_chart_series(product, index_data, "unit", start_date, end_date)
        cum_points, cum_index_points, cum_note = self.build_chart_series(product, index_data, "cum", start_date, end_date)
        short_product_name = product["fund_name"][:18]
        short_index_name = index_name[:14]

        self.draw_line_chart(
            self.unit_chart_canvas,
            "单位净值与指数换算值",
            unit_points,
            unit_index_points,
            "产品单位净值",
            f"{short_index_name}换算值",
            unit_note or "单位净值图暂无可绘制数据",
        )
        self.draw_line_chart(
            self.cum_chart_canvas,
            "累计净值与指数换算值",
            cum_points,
            cum_index_points,
            "产品累计净值",
            f"{short_index_name}换算值",
            cum_note or "累计净值图暂无可绘制数据",
        )
        self.chart_note_label.config(
            text=f"当前产品：{short_product_name}；指数：{index_name}；{unit_note or cum_note}"
        )

    def upload_file(self):
        filepaths = filedialog.askopenfilenames(
            title="选择一个或多个净值 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")]
        )
        if not filepaths:
            return

        loaded_products = []
        errors = []
        for filepath in filepaths:
            try:
                fund_name, data = load_fund_data(filepath)
                loaded_products.append({
                    "fund_name": fund_name,
                    "filepath": filepath,
                    "filename": Path(filepath).name,
                    "data": data,
                    "one_year_data": None,
                    "results": [],
                })
            except Exception as exc:
                errors.append(f"{Path(filepath).name}: {exc}")

        if not loaded_products:
            messagebox.showerror("错误", "没有成功读取任何文件。\n\n" + "\n".join(errors[:5]))
            return

        self.products = loaded_products
        self.ensure_product_ids()
        self.selected_index = 0
        self.update_product_list()
        self.update_weekly_date_options()
        self.update_display()
        self.set_export_button_state(True)
        self.file_label.config(
            text=(
                f"已读取 {len(self.products)} 个产品 / {len(filepaths)} 个文件\n"
                f"当前：{self.products[0]['filename']}\n"
                f"{self.products[0]['data'][0][0]} 至 {self.products[0]['data'][-1][0]}"
            )
        )
        if len(self.products) > 1:
            self.notebook.select(self.compare_tab)

        if errors:
            messagebox.showwarning(
                "部分文件读取失败",
                f"成功读取 {len(self.products)} 个文件，失败 {len(errors)} 个。\n\n" + "\n".join(errors[:8])
            )

    def ensure_product_ids(self):
        for product in self.products:
            if not product.get("product_id"):
                product["product_id"] = f"p{self.next_product_id}"
                self.next_product_id += 1

    def update_product_list(self):
        self.ensure_product_ids()
        self.product_listbox.delete(0, tk.END)
        for idx, product in enumerate(self.products, start=1):
            label = f"{idx}. {product['fund_name']} ({product['filename']})"
            self.product_listbox.insert(tk.END, label)
        if self.products and self.selected_index is not None:
            self.product_listbox.selection_clear(0, tk.END)
            self.product_listbox.selection_set(self.selected_index)
            self.product_listbox.activate(self.selected_index)

    def move_selected_product(self, direction):
        if not self.products:
            return

        selection = self.product_listbox.curselection()
        current_index = selection[0] if selection else (self.selected_index or 0)
        new_index = current_index + direction
        if new_index < 0 or new_index >= len(self.products):
            return

        self.products[current_index], self.products[new_index] = self.products[new_index], self.products[current_index]
        self.selected_index = new_index
        self.update_product_list()
        self.update_display()

    def on_product_list_press(self, event):
        if not self.products:
            self.dragged_product_index = None
            return
        index = self.product_listbox.nearest(event.y)
        if 0 <= index < len(self.products):
            self.dragged_product_index = index
            self.product_listbox.selection_clear(0, tk.END)
            self.product_listbox.selection_set(index)
            self.product_listbox.activate(index)
            self.selected_index = index

    def on_product_list_drag(self, event):
        if self.dragged_product_index is None:
            return
        target_index = self.product_listbox.nearest(event.y)
        if target_index == self.dragged_product_index or target_index < 0 or target_index >= len(self.products):
            return

        product = self.products.pop(self.dragged_product_index)
        self.products.insert(target_index, product)
        self.dragged_product_index = target_index
        self.selected_index = target_index
        self.update_product_list()

    def on_product_list_release(self, _event):
        if self.dragged_product_index is None:
            return
        self.dragged_product_index = None
        self.update_display()

    def on_product_select(self, _event=None):
        selection = self.product_listbox.curselection()
        if not selection:
            return
        current_tab = self.notebook.select()
        self.selected_index = selection[0]
        self.update_display()
        if current_tab == str(self.chart_tab):
            self.notebook.select(self.chart_tab)
        else:
            self.notebook.select(self.detail_tab)

    def set_export_button_state(self, enabled):
        if enabled:
            self.export_btn.config(
                state="normal",
                bg=self.colors["orange"],
                activebackground="#5f220d",
                cursor="hand2",
            )
        else:
            self.export_btn.config(
                state="disabled",
                bg=self.colors["disabled"],
                activebackground=self.colors["disabled"],
                cursor="arrow",
            )

    def get_selected_product(self):
        if not self.products:
            return None
        if self.selected_index is None or self.selected_index >= len(self.products):
            self.selected_index = 0
        return self.products[self.selected_index]

    def parse_settings(self):
        try:
            self.rf_annual = float(self.rf_entry.get())
        except ValueError as exc:
            raise ValueError("无风险利率必须填写数字，例如 1.82") from exc
        self.use_prev_year_end = self.rf_var.get()

    def recalculate_products(self):
        self.parse_settings()
        for product in self.products:
            product["one_year_data"] = calc_one_year_return(product["data"])
            product["results"] = build_report(
                product["fund_name"],
                product["data"],
                self.rf_annual,
                self.use_prev_year_end,
            )

    def sync_current_product(self, product):
        self.fund_name = product["fund_name"]
        self.filepath = product["filepath"]
        self.data = product["data"]
        self.one_year_data = product.get("one_year_data")
        self.results = product.get("results", [])

    def format_nav(self, value):
        return "--" if value is None else f"{value:.4f}"

    def format_percent(self, value, signed=True):
        if value is None:
            return "--"
        return f"{value:+.2f}" if signed else f"{value:.2f}"

    def update_display(self):
        if not self.products:
            return

        self.ensure_product_ids()
        self.recalculate_products()
        product = self.get_selected_product()
        if not product:
            return
        self.sync_current_product(product)

        latest = self.data[-1]

        self.product_name_label.config(text=self.fund_name)
        self.product_meta_label.config(
            text=(
                f"这是产品“{self.fund_name}”的净值分析结果。"
                f"当前为第 {self.selected_index + 1} / {len(self.products)} 个产品，"
                f"读取 {len(self.data)} 条净值记录，最新净值日期为 {latest[0]}。"
            )
        )
        self.summary_labels["file"].config(text=Path(self.filepath).name if self.filepath else "--")
        self.summary_labels["period"].config(text=f"{self.data[0][0]} 至 {self.data[-1][0]}")
        self.summary_labels["count"].config(text=f"{len(self.data)} 条")
        self.summary_labels["rf"].config(text=f"{self.rf_annual:.2f}%")

        self.metric_value_labels["latest_nav"].config(text=f"{latest[1]:.4f}", fg=self.colors["text"])
        self.metric_hint_labels["latest_nav"].config(text="单位净值，来自最新一条净值记录")

        if latest[2] is not None:
            self.metric_value_labels["cum_nav"].config(text=f"{latest[2]:.4f}", fg=self.colors["text"])
            self.metric_hint_labels["cum_nav"].config(text="累计单位净值，来自累计净值列")
        else:
            self.metric_value_labels["cum_nav"].config(text="无", fg=self.colors["muted"])
            self.metric_hint_labels["cum_nav"].config(text="文件中未识别到累计净值列")

        self.metric_value_labels["latest_date"].config(text=str(latest[0]), fg=self.colors["text"])
        self.metric_hint_labels["latest_date"].config(text=f"数据起点：{self.data[0][0]}")

        if self.one_year_data:
            one_year_return = self.one_year_data["one_year_return"]
            return_color = self.colors["red"] if one_year_return >= 0 else self.colors["green"]
            self.metric_value_labels["one_year"].config(text=f"{one_year_return:+.2f}%", fg=return_color)
            self.metric_hint_labels["one_year"].config(
                text=(
                    f"{self.one_year_data['start_date']} "
                    f"{self.one_year_data['start_nav']:.4f} 至 "
                    f"{self.one_year_data['end_nav']:.4f}"
                )
            )
        else:
            self.metric_value_labels["one_year"].config(text="数据不足", fg=self.colors["muted"])
            self.metric_hint_labels["one_year"].config(text="净值历史不足以形成区间对比")

        for item in self.tree.get_children():
            self.tree.delete(item)

        basis_text = "自然年度使用前一年 12-31 净值作年初基准" if self.use_prev_year_end else "自然年度使用当年首条净值作年初基准"
        self.table_note_label.config(text=f"含近1/2/3年、自然年度、成立以来；{basis_text}")

        for m in self.results:
            sign_return = f"{m['total_return']:+.2f}" if m['total_return'] >= 0 else f"{m['total_return']:.2f}"
            sign_ann = f"{m['ann_return']:+.2f}" if m['ann_return'] >= 0 else f"{m['ann_return']:.2f}"
            cum_nav_text = f"{m['cum_nav']:.4f}" if m.get("cum_nav") is not None else "--"
            values = [
                m["label"],
                str(m["start_date"]),
                str(m["end_date"]),
                f"{m['start_nav']:.4f}",
                f"{m['end_nav']:.4f}",
                cum_nav_text,
                sign_return,
                sign_ann,
                f"{m['max_drawdown']:.2f}",
                f"{m['ann_volatility']:.2f}",
                f"{m['sharpe_ratio']:.4f}",
                f"{m['sortino_ratio']:.4f}",
                f"{m['calmar_ratio']:.4f}",
                m.get("note", ""),
            ]
            row_tag = "positive" if m["total_return"] >= 0 else "negative"
            self.tree.insert("", "end", values=values, tags=(row_tag,))

        self.update_compare_display()
        self.update_weekly_display()
        self.update_chart_date_options()
        self.update_chart_display()
        self.file_label.config(
            text=(
                f"已读取 {len(self.products)} 个产品\n"
                f"当前：{product['filename']}\n"
                f"{self.data[0][0]} 至 {self.data[-1][0]}"
            )
        )

    def update_compare_display(self):
        for item in self.compare_tree.get_children():
            self.compare_tree.delete(item)

        if not self.products:
            self.compare_note_label.config(text="批量上传后显示所有产品核心指标")
            return

        self.compare_note_label.config(
            text=f"共 {len(self.products)} 个产品；近1/2/3年不足区间会按可用数据计算并在备注说明"
        )
        for product in self.products:
            row = get_product_compare_values(product)
            values = [
                row["fund"],
                row["file"],
                str(row["start_date"]),
                str(row["latest_date"]),
                row["records"],
                self.format_nav(row["latest_nav"]),
                self.format_nav(row["latest_cum_nav"]),
                self.format_percent(row["near_one_return"]),
                self.format_percent(row["near_two_return"]),
                self.format_percent(row["near_three_return"]),
                self.format_percent(row["since_return"]),
                self.format_percent(row["since_max_drawdown"], signed=False),
                self.format_nav(row["since_sharpe"]),
                row["note"],
            ]
            self.compare_tree.insert("", "end", iid=product.get("product_id"), values=values)

    def update_weekly_date_options(self):
        dates = sorted({record[0] for product in self.products for record in product["data"]})
        self.weekly_date_options = dates
        values = [str(d) for d in dates]

        for combo in self.weekly_date_boxes.values():
            combo.configure(values=values)

        if not values:
            for var in self.weekly_date_vars.values():
                var.set("")
            return

        latest = values[-1]
        previous = values[-2] if len(values) >= 2 else latest
        defaults = {
            "unit": previous,
            "prev": previous,
            "current": latest,
            "cum": latest,
        }
        for key, default_value in defaults.items():
            current_value = self.weekly_date_vars[key].get()
            if current_value not in values:
                self.weekly_date_vars[key].set(default_value)

    def get_weekly_selected_dates(self):
        dates = {}
        for key, var in self.weekly_date_vars.items():
            selected_date = parse_date(var.get())
            if selected_date is None:
                return None
            dates[key] = selected_date
        return dates

    def get_benchmark_rows(self, dates=None):
        benchmark_rows = []
        invalid_inputs = []
        weekly_period = "--"
        interval_period = "--"
        if dates:
            weekly_period = f"{dates['prev']} 至 {dates['current']}"
            interval_period = f"{dates['unit']} 至 {dates['current']}"

        for name in BENCHMARK_NAMES:
            vars_for_name = self.benchmark_vars.get(name, {})
            prices = {}
            labels = {
                "start": "区间起点收盘",
                "prev": "上周收盘",
                "current": "本周收盘",
            }
            for key, label in labels.items():
                text = vars_for_name[key].get() if key in vars_for_name else ""
                try:
                    prices[key] = parse_percent_input(text)
                except ValueError:
                    prices[key] = None
                    invalid_inputs.append(f"{name} {label}")

            weekly_return = calc_return_from_prices(prices["prev"], prices["current"])
            interval_return = calc_return_from_prices(prices["start"], prices["current"])

            benchmark_rows.append({
                "name": name,
                "weekly_period": weekly_period,
                "start_close": prices["start"],
                "prev_close": prices["prev"],
                "current_close": prices["current"],
                "weekly_return": weekly_return,
                "interval_period": interval_period,
                "interval_return": interval_return,
            })

        return benchmark_rows, invalid_inputs

    def make_benchmark_result_text(self, weekly_return, interval_return, benchmark_rows):
        parts = []
        for benchmark in benchmark_rows:
            subparts = []
            weekly_benchmark = benchmark.get("weekly_return")
            interval_benchmark = benchmark.get("interval_return")
            if weekly_return is not None and weekly_benchmark is not None:
                diff = weekly_return - weekly_benchmark
                status = "跑赢" if diff >= 0 else "未跑赢"
                subparts.append(f"本周{status}{diff:+.2f}pct")
            if interval_return is not None and interval_benchmark is not None:
                diff = interval_return - interval_benchmark
                status = "跑赢" if diff >= 0 else "未跑赢"
                subparts.append(f"区间{status}{diff:+.2f}pct")
            if subparts:
                parts.append(f"{benchmark['name']}: {' / '.join(subparts)}")
        return "；".join(parts) if parts else "请填写指数收盘价"

    def get_weekly_product_rows(self):
        if not self.products:
            return [], []

        self.ensure_product_ids()
        dates = self.get_weekly_selected_dates()
        if not dates:
            return [], []

        benchmark_rows, invalid_inputs = self.get_benchmark_rows(dates)
        rows = []
        for product in self.products:
            data = product["data"]
            unit_record = get_nav_on_or_before(data, dates["unit"])
            prev_record = get_nav_on_or_before(data, dates["prev"])
            current_record = get_nav_on_or_before(data, dates["current"])
            cum_record = get_nav_on_or_before(data, dates["cum"])
            note_parts = []

            date_records = [
                ("单位净值", "unit", unit_record),
                ("上周收盘", "prev", prev_record),
                ("本周收盘", "current", current_record),
                ("累计净值", "cum", cum_record),
            ]
            for label, key, record in date_records:
                if record is None:
                    note_parts.append(f"{label}无可用净值")
                elif record[0] != dates[key]:
                    note_parts.append(f"{label}用{record[0]}")

            weekly_return = None
            if prev_record and current_record and current_record[0] > prev_record[0] and prev_record[1] > 0:
                weekly_return = (current_record[1] / prev_record[1] - 1) * 100
            elif prev_record and current_record:
                note_parts.append("本周区间日期不足")

            interval_return = None
            if unit_record and current_record and current_record[0] > unit_record[0] and unit_record[1] > 0:
                interval_return = (current_record[1] / unit_record[1] - 1) * 100
            elif unit_record and current_record:
                note_parts.append("区间收益日期不足")

            cum_nav = cum_record[2] if cum_record and len(cum_record) > 2 else None
            if cum_record and cum_nav is None:
                note_parts.append("累计净值无数据")

            rows.append({
                "product_id": product["product_id"],
                "fund": product["fund_name"],
                "unit_date": str(unit_record[0]) if unit_record else "--",
                "unit_nav": unit_record[1] if unit_record else None,
                "prev_date": str(prev_record[0]) if prev_record else "--",
                "prev_nav": prev_record[1] if prev_record else None,
                "current_date": str(current_record[0]) if current_record else "--",
                "current_nav": current_record[1] if current_record else None,
                "cum_date": str(cum_record[0]) if cum_record else "--",
                "cum_nav": cum_nav,
                "weekly_return": weekly_return,
                "interval_return": interval_return,
                "benchmark_result": self.make_benchmark_result_text(weekly_return, interval_return, benchmark_rows),
                "note": "；".join(dict.fromkeys(note_parts)),
            })

        return rows, benchmark_rows

    def update_weekly_display(self, _event=None):
        for item in self.weekly_tree.get_children():
            self.weekly_tree.delete(item)
        for item in self.benchmark_tree.get_children():
            self.benchmark_tree.delete(item)

        if not self.products:
            self.weekly_note_label.config(text="上传净值文件后可生成周报涨幅")
            return

        dates = self.get_weekly_selected_dates()
        if not dates:
            self.weekly_note_label.config(text="请先选择完整的周报日期")
            return

        benchmark_rows, invalid_inputs = self.get_benchmark_rows(dates)
        for benchmark in benchmark_rows:
            self.benchmark_tree.insert("", "end", values=[
                benchmark["name"],
                self.format_nav(benchmark["start_close"]),
                self.format_nav(benchmark["prev_close"]),
                self.format_nav(benchmark["current_close"]),
                self.format_percent(benchmark["weekly_return"]),
                self.format_percent(benchmark["interval_return"]),
            ])

        rows, benchmark_rows = self.get_weekly_product_rows()
        for row in rows:
            values = [
                row["fund"],
                row["unit_date"],
                self.format_nav(row["unit_nav"]),
                row["prev_date"],
                self.format_nav(row["prev_nav"]),
                row["current_date"],
                self.format_nav(row["current_nav"]),
                row["cum_date"],
                self.format_nav(row["cum_nav"]),
                self.format_percent(row["weekly_return"]),
                self.format_percent(row["interval_return"]),
                row["benchmark_result"],
                row["note"],
            ]
            row_tag = "positive" if (row["weekly_return"] or 0) >= 0 else "negative"
            self.weekly_tree.insert("", "end", iid=row["product_id"], values=values, tags=(row_tag,))

        note = (
            f"共 {len(rows)} 个产品；本周涨幅 = 本周收盘单位净值 / 上周收盘单位净值 - 1；"
            "区间收益率 = 本周收盘单位净值 / 单位净值日期净值 - 1；指数涨幅根据收盘价自动计算"
        )
        if invalid_inputs:
            note += f"；请检查输入：{', '.join(invalid_inputs)}"
        self.weekly_note_label.config(text=note)

    def recalculate(self):
        if self.products:
            try:
                self.update_display()
            except Exception as e:
                messagebox.showerror("错误", f"重新计算失败：{str(e)}")

    def export_results(self):
        if not self.products:
            messagebox.showwarning("警告", "请先加载数据")
            return

        try:
            self.update_display()
        except Exception as e:
            messagebox.showerror("错误", f"导出前计算失败：{str(e)}")
            return

        filepath = filedialog.asksaveasfilename(
            title="导出结果",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile="多产品净值分析结果.xlsx" if len(self.products) > 1 else f"{self.fund_name}_分析结果.xlsx"
        )
        if filepath:
            try:
                weekly_rows, benchmark_rows = self.get_weekly_product_rows()
                save_all_products_workbook(filepath, self.products, weekly_rows, benchmark_rows)
                messagebox.showinfo("成功", f"结果已导出至：\n{filepath}")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败：{str(e)}")


def main():
    root = tk.Tk()
    app = FundAnalyzerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
